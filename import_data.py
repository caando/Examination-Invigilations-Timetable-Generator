import datetime
from classes import db, Teacher, Paper, Lesson, Fixed_Invigilation
import xmltodict
from openpyxl import load_workbook
from sqlalchemy.orm import noload

# These are sample past year mid year papers test cases

def import_paper_data(filename):
	for paper in Paper.query.all():
		db.session.delete(paper)
	for invigilation in Fixed_Invigilation.query.all():
		db.session.delete(invigilation)
	for ws in load_workbook(filename).worksheets:
		row = 2
		while ws.cell(row, 1).value:
			subj = ws.cell(row, 1).value
			name = ws.cell(row, 2).value
			dept = ws.cell(row, 3).value
			count = int(ws.cell(row, 4).value)//40 + int(int(ws.cell(row, 4).value) % 40 > 0) + 1
			date = ws.cell(row, 5).value.date()
			starttime = ws.cell(row, 6).value
			endtime = ws.cell(row, 7).value
			level = int(ws.cell(row, 8).value)
			new_paper = Paper(subj, name, dept, count, date, starttime, endtime, level)
			db.session.add(new_paper)
			for teacher in Teacher.query.options(noload("Invigilations")).all():
				if ws.cell(row, 9).value != None:
					if teacher.Name in ws.cell(row, 9).value:
						Fixed_Invigilation(teacher, new_paper, coordinator=True)
				if ws.cell(row, 10).value != None:
					if teacher.Name in ws.cell(row, 10).value:
						Fixed_Invigilation(teacher, new_paper, coordinator=True)
			row += 1
	db.session.commit()

def strip_parent(string):
	if len(string) > 0:
		if string[0] in ' ()':
			string = strip_parent(string[1:])
		if string[-1] in ' ()':
			string = strip_parent(string[:-1])
	return string

def shorten_dept(dept):
	if '(' in strip_parent(dept):
		return strip_parent(dept).split('(')[1]
	elif '-' in strip_parent(dept):
		return strip_parent(dept).split('-')[1]
	elif ',' in strip_parent(dept):
		return strip_parent(dept).split(',')[1]
	elif ' ' in strip_parent(dept):
		return strip_parent(dept).split(' ')[1]
	return strip_parent(dept)

def get_load(info):
	if "HOD" in info:
		return 2/3
	related_to_load = False
	load_str = ''
	for char in strip_parent(info):
		if char.isdigit() or char == '/':
			load_str += char
			if char.isdigit():
				related_to_load = True
	if related_to_load:
		return float(load_str[0])/float(load_str[-1])
	return 1.0

def import_teacher_data(string):
	dept_str = ''
	fullname = ''
	info_str = ''
	parent_level = 0
	dept_comp = False
	for i in range(len(string)):
		if string[i] == '(':
			parent_level += 1
		if string[i] == ')':
			parent_level -= 1
			if dept_comp == False and parent_level == 0:
				dept_comp = True
		if parent_level > 0:
			if dept_comp:
				info_str += string[i]
			else:
				dept_str += string[i]
		if parent_level == 0:
			fullname += string[i]
	existing_teacher = Teacher.query.filter(Teacher.Name == strip_parent(fullname)).first()
	if existing_teacher:
		return existing_teacher
	else:
		new_teacher = Teacher(strip_parent(fullname), shorten_dept(dept_str), get_load(info_str))
		db.session.add(new_teacher)
		return new_teacher

days = {'11111':0, '10000':1, '01000':2, '00100':3, '00010':4, '00001':5,}

class lesson_temp():
	def __init__(self, subject_name):
		self.title = subject_name
		self.teachers = []
		self.level = None
		self.periods = [{}, {}, {}, {}, {}]
	def add_teacher(self, teacher):
		self.teachers.append(teacher)
	def set_level(self, level):
		self.level = level
	def add_period(self, name, period, day):
		self.periods[day-1][name] = period
	def set_start_n_end(self, day):
		self.starttime = datetime.datetime.strptime(self.periods[day-1][min(k for k in self.periods[day-1].keys())][0], '%H:%M').time()
		if max(k for k in self.periods[day-1].keys()) == '31':
			self.endtime = datetime.time(18, 0, 0)
		else:
			self.endtime = datetime.datetime.strptime(self.periods[day-1][max(k for k in self.periods[day-1].keys())][1], '%H:%M').time()

# Import from xml file
def import_to_db(filename):
	for teacher in Teacher.query.all():
		db.session.delete(teacher)
	for lesson in Lesson.query.all():
		db.session.delete(lesson)
	with open(filename, 'r') as rf:
		doc = xmltodict.parse(rf.read())
		teachers = {}
		periods = {}
		subjects = {}
		lessons = {}
		classes = {}
		for Class in doc['timetable']['classes']['class']:
			if Class['@name'][0:2] == '18':
				classes[Class['@id']] = 6
			elif Class['@name'][0:2] == '19':
				classes[Class['@id']] = 5
			elif Class['@name'][0] == '1':
				classes[Class['@id']] = 1
			elif Class['@name'][0] == '2':
				classes[Class['@id']] = 2
			elif Class['@name'][0] == '3':
				classes[Class['@id']] = 3
			elif Class['@name'][0] == '4':
				classes[Class['@id']] = 4
		for teacher in doc['timetable']['teachers']['teacher']:
			if teacher['@name']:
				teachers[teacher['@id']] = import_teacher_data(teacher['@name'])
		db.session.commit()
		for period in doc['timetable']["periods"]['period']:
			periods[period['@period']] = [period['@starttime'], period['@endtime']]
		for subject in doc['timetable']["subjects"]['subject']:
			subjects[subject['@id']] = subject['@name']
		for lesson in doc['timetable']["lessons"]['lesson']:
			lessons[lesson['@id']] = lesson_temp(subjects[lesson['@subjectid']])
			if lesson['@teacherids']:
				teacherids = lesson['@teacherids'].split(',')
				for teacherid in teacherids:
					lessons[lesson['@id']].add_teacher(teachers[teacherid])
			if lesson['@classids']:
				classesid = lesson['@classids'].split(',')
				level = classes[classesid[0]]
				for Classid in classesid:
					if classes[Classid] != level:
						level = None
						break
				lessons[lesson['@id']].set_level(level)
		for card in doc['timetable']["cards"]['card']:
			if days[card['@days']] == 0:
				for i in range(1, 6):
					lessons[card['@lessonid']].add_period(card['@period'], periods[card['@period']], i)
			else:
				lessons[card['@lessonid']].add_period(card['@period'], periods[card['@period']], days[card['@days']])
		for lesson in lessons.values():
			for i in range(len(lesson.periods)):
				if lesson.periods[i]:
					lesson.set_start_n_end(i+1)
					if lesson.teachers:
						for teacher in lesson.teachers:
							new_lesson = Lesson(lesson.title, i + 1, lesson.starttime, lesson.endtime, teacher)
							if lesson.level:
								new_lesson.set_level(lesson.level)
	db.session.commit()
