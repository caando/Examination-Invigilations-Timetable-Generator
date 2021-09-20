from classes import db, Teacher, Paper, Invigilation, Version, Pool, Fixed_Invigilation
import datetime
from sqlalchemy.orm import noload
from openpyxl import Workbook

cohort = ["Year 1", "Year 2", "Year 3", "Year 4", "JC 1", "JC 2"]

def export_analytics():
	versions = Version.query.options(noload("Invigilations")).all()
	wb = Workbook()
	wb.remove_sheet(wb.get_sheet_by_name(wb.get_sheet_names()[0]))
	ws = wb.create_sheet("Analytics")

	for i in range(len(versions)//100):
		total = 0
		for j in range(100):
			total += versions[i*100 + j].Score
		ws.cell(row=i+1, column=1, value = total/100)

	wb.save('Analytics.xlsx')

def export_teachers():
	versions = Version.query.all()
	versions.sort(key=lambda version:version.Score)

	papers = Paper.query.options(noload('Invigilators')).all()
	papers.sort(key=lambda paper:paper.Date)

	teachers = Teacher.query.options(noload('Invigilations')).all()
	teachers.sort(key=lambda teacher:teacher.Department)
	def output_excel(wb, version):
		ws = wb.create_sheet(str(version.Score))
		ws.cell(row=2, column=1, value = "Name")
		ws.cell(row=2, column=2, value = "Department")
		ws.cell(row=2, column=3, value = "Load")
		ws.cell(row=2, column=4, value = "Remarks")
		ws.cell(row=2, column=5, value = "Reserve")
		ws.cell(row=2, column=6, value = "Hours")
		ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
		dates = []
		for i in range(len(papers)):
			ws.cell(row=2, column=7+i, value = str(papers[i].Name))
			if i != 0:
				if papers[i].Date != papers[i-1].Date:
					dates.append([papers[i].Date, 1])
				else:
					dates[-1][1] += 1
			else:
				dates.append([papers[i].Date, 1])
		start_date = 7
		for date in dates:
			ws.merge_cells(start_row=1, start_column=start_date, end_row=1, end_column=start_date + date[1]-1)
			ws.cell(row=1, column=start_date, value = str(date[0]))
			start_date += date[1]
		teachers_dictionary = {}
		for teacher in teachers:
			teachers_dictionary[teacher.id] = teacher
			teachers_dictionary[teacher.id].invigilations = []
			teachers_dictionary[teacher.id].hours = 0
		for invigilation in version.Invigilations:
			paper = Paper.query.options(noload('Invigilators')).get(invigilation.PaperId)
			teachers_dictionary[invigilation.TeacherId].invigilations.append(paper.id)
			teachers_dictionary[invigilation.TeacherId].hours += paper.EndTime.hour - paper.StartTime.hour + (paper.EndTime.minute - paper.StartTime.minute)/60
		for invigilation in Fixed_Invigilation.query.all():
			paper = Paper.query.options(noload('Invigilators')).get(invigilation.PaperId)
			teachers_dictionary[invigilation.TeacherId].invigilations.append(paper.id)
			teachers_dictionary[invigilation.TeacherId].hours += paper.EndTime.hour - paper.StartTime.hour + (paper.EndTime.minute - paper.StartTime.minute)/60
		for i in range(len(teachers)):
			ws.cell(row=3 + i, column= 1, value = teachers[i].Name)
			ws.cell(row=3 + i, column= 2, value = teachers[i].Department)
			ws.cell(row=3 + i, column= 3, value = teachers[i].Load)
			ws.cell(row=3 + i, column= 6, value = teachers[i].hours)
			for j in range(len(papers)):
				if papers[j].id in teachers[i].invigilations:
					ws.cell(row=3 + i, column=7 + j, value = 1)
	wb = Workbook()
	wb.remove_sheet(wb.worksheets[0])
	for version in versions[:10]:
		output_excel(wb, version)
	wb.save('output_teachers.xlsx')


def export_papers():
	versions = Version.query.all()
	versions.sort(key=lambda version:version.Score)

	papers = Paper.query.options(noload('Invigilators')).all()
	papers.sort(key=lambda paper:paper.Date)

	teachers = Teacher.query.options(noload('Invigilations')).all()
	teachers.sort(key=lambda teacher:teacher.Department)
	def output_excel(wb, version):
		ws = wb.create_sheet(str(version.Score))
		ws.cell(row=1, column=1, value = "Date")
		ws.cell(row=1, column=2, value = "Cohort")
		ws.cell(row=1, column=3, value = "Venue")
		ws.cell(row=1, column=4, value = "Start Time")
		ws.cell(row=1, column=5, value = "End Time")
		ws.cell(row=1, column=6, value = "Paper")
		ws.cell(row=1, column=7, value = "Coordinator")
		ws.cell(row=1, column=8, value = "Department")
		ws.cell(row=1, column=9, value = "Invigilators")
		ws.cell(row=1, column=10, value = "Department")
		dictionary = {}
		row = 2
		for invigilation in version.Invigilations:
			if invigilation.PaperId in dictionary.keys():
				dictionary[invigilation.PaperId].append(invigilation.TeacherId)
			else:
				dictionary[invigilation.PaperId] = [invigilation.TeacherId]
		for paper in papers:
			if paper != papers[0]:
				if paper.Date != papers[papers.index(paper)-1].Date:
					ws.cell(row=row, column=1, value = paper.Date)
			else:
				ws.cell(row=row, column=1, value = paper.Date)
			ws.cell(row=row, column=2, value = cohort[paper.Level-1])
			ws.cell(row=row, column=4, value = paper.StartTime)
			ws.cell(row=row, column=5, value = paper.EndTime)
			ws.cell(row=row, column=6, value = paper.Name)
			coordinators = []
			normal = []
			for i in range(len(paper.Fixed_Invigilators)):
				if paper.Fixed_Invigilators[i].coordinator:
					coordinators.append(paper.Fixed_Invigilators[i].TeacherId)
				else:
					normal.append(paper.Fixed_Invigilators[i].TeacherId)
			if paper.id in dictionary.keys():
				for Id in dictionary[paper.id]:
					normal.append(Id)
			for i in range(len(coordinators)):
				ws.cell(row = row + i, column=7, value = Teacher.query.options(noload('Invigilations')).get(coordinators[i]).Name)
				ws.cell(row = row + i, column=8, value = Teacher.query.options(noload('Invigilations')).get(coordinators[i]).Department)
			for i in range(len(normal)):
				ws.cell(row = row + i, column=9, value = Teacher.query.options(noload('Invigilations')).get(normal[i]).Name)
				ws.cell(row = row + i, column=10, value = Teacher.query.options(noload('Invigilations')).get(normal[i]).Department)
			row = max(row + len(coordinators), row + len(normal))
	wb = Workbook()
	wb.remove_sheet(wb.worksheets[0])
	for version in versions[:10]:
		output_excel(wb, version)
	wb.save('output_papers.xlsx')

